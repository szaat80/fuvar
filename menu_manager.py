from PySide6.QtWidgets import QMenuBar, QMenu, QMessageBox, QDialog
from PySide6.QtCore import QTimer
from openpyxl import load_workbook
import os

class MenuBar(QMenuBar):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.createFileMenu()
        self.createDatabaseMenu()

    def createFileMenu(self):
        fileMenu = self.addMenu("Fájl")
        fileMenu.addAction("Excel megnyitása").triggered.connect(self.openExcel)
        fileMenu.addAction("Mentés").triggered.connect(self.saveWorkHours)
        fileMenu.addAction("Fuvar mentés").triggered.connect(self.saveDelivery)
        fileMenu.addAction("Nyomtatás").triggered.connect(self.printData)
        fileMenu.addAction("Kilépés").triggered.connect(self.parent().close)

    def createDatabaseMenu(self):
        dbMenu = self.addMenu("Adatbázis")
        dbMenu.addAction("Törzsadatok kezelése").triggered.connect(self.openDatabaseManager)

    def openExcel(self):
        try:
            if os.path.exists('munkaora_nyilvantartas.xlsx'):
                os.startfile('munkaora_nyilvantartas.xlsx')
                QTimer.singleShot(2000, self.loadFromExcel)
            else:
                QMessageBox.warning(self.parent(), "Figyelmeztetés", "Excel fájl nem található!")
        except Exception as e:
            QMessageBox.warning(self.parent(), "Hiba", f"Excel megnyitási hiba: {str(e)}")

    def loadFromExcel(self):
        try:
            if not os.path.exists('munkaora_nyilvantartas.xlsx'):
                return
                
            wb = load_workbook('munkaora_nyilvantartas.xlsx')
            ws = wb.active
            
            # Clear existing data
            work_table = self.parent().work_table
            for row in range(work_table.rowCount()):
                for col in range(2, work_table.columnCount()):
                    work_table.setItem(row, col, None)
            
            # Load new data
            for excel_row in range(2, ws.max_row + 1):
                date = ws.cell(row=excel_row, column=1).value
                if not date:
                    continue
                    
                for table_row in range(work_table.rowCount()):
                    if work_table.item(table_row, 0).text() == str(date):
                        for col in range(2, work_table.columnCount()):
                            value = ws.cell(row=excel_row, column=col).value
                            if value is not None:
                                item = QTableWidgetItem(str(value))
                                item.setTextAlignment(Qt.AlignCenter)
                                work_table.setItem(table_row, col-1, item)
                        break
                
            QMessageBox.information(self.parent(), "Siker", "Excel adatok betöltve!")
            
        except Exception as e:
            QMessageBox.warning(self.parent(), "Hiba", f"Excel betöltési hiba: {str(e)}")

    def saveWorkHours(self):
        self.parent().work_hours_manager.saveWorkHoursToExcel()

    def saveDelivery(self):
        self.parent().delivery_manager.saveDeliveryData()

    def openDatabaseManager(self):
        try:
            dbManager = self.parent().database_manager
            dbManager.exec()
        except Exception as e:
            QMessageBox.critical(self.parent(), "Hiba", f"Hiba a törzsadat kezelő megnyitásakor: {str(e)}")

    def printData(self):
        dialog = QPrintDialog(self.parent())
        if dialog.exec() == QDialog.Accepted:
            printer = dialog.printer()
            # TODO: Implement printing logic