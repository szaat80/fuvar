from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QMessageBox
from PySide6.QtCore import Qt, QDate, QTime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os
from datetime import datetime

class WorkHoursManager:
    def __init__(self, parent=None):
        self.parent = parent
        self.work_table = None

    def setup_work_table(self, table):
        self.work_table = table
        self.setup_headers()
        
    def setup_headers(self):
        headers = [
            "Dátum", "Nap", 
            "Sima Munkanap\nKezdés", "Sima Munkanap\nVégzés", "Ledolgozott\nÓrák",
            "Műhely\nKezdés", "Műhely\nVégzés", "Ledolgozott\nÓrák",
            "Szabadság", "Betegszabadság\n(TP)"
        ]
        self.work_table.setColumnCount(len(headers))
        self.work_table.setHorizontalHeaderLabels(headers)

    def saveWorkHours(self):
        try:
            current_date = self.parent.date_edit.date().toString('yyyy-MM-dd')
            work_type = self.parent.type_combo.currentText()
            current_driver = self.parent.driver_combo.currentText()
            month_dir = os.path.join('driver_records', current_driver, 
                                    f"{datetime.now().year}_{datetime.now().month:02d}")
            os.makedirs(month_dir, exist_ok=True)
            excel_path = os.path.join(month_dir, 'work_hours.xlsx')

            row_to_update = -1
            for row in range(self.work_table.rowCount()):
                if self.work_table.item(row, 0).text() == current_date:
                    row_to_update = row
                    break

            if row_to_update == -1:
                raise Exception("Dátum nem található")

            if work_type in ["Szabadság", "Betegszabadság (TP)"]:
                col_index = 8 if work_type == "Szabadság" else 9
                item = QTableWidgetItem("1")
                item.setTextAlignment(Qt.AlignCenter)
                self.work_table.setItem(row_to_update, col_index, item)
            else:
                start_time = self.parent.start_time.time().toString('HH:mm')
                end_time = self.parent.end_time.time().toString('HH:mm')
                
                if work_type == "Sima munkanap":
                    self.work_table.setItem(row_to_update, 2, QTableWidgetItem(start_time))
                    self.work_table.setItem(row_to_update, 3, QTableWidgetItem(end_time))
                    start = QTime.fromString(start_time, 'HH:mm')
                    end = QTime.fromString(end_time, 'HH:mm')
                    if start.isValid() and end.isValid():
                        hours = round(start.secsTo(end) / 3600.0, 2)
                        self.work_table.setItem(row_to_update, 4, QTableWidgetItem(f"{hours:.2f}"))
                else:  # Műhely nap
                    self.work_table.setItem(row_to_update, 5, QTableWidgetItem(start_time))
                    self.work_table.setItem(row_to_update, 6, QTableWidgetItem(end_time))
                    start = QTime.fromString(start_time, 'HH:mm')
                    end = QTime.fromString(end_time, 'HH:mm')
                    if start.isValid() and end.isValid():
                        hours = round(start.secsTo(end) / 3600.0, 2)
                        self.work_table.setItem(row_to_update, 7, QTableWidgetItem(f"{hours:.2f}"))

            wb = Workbook()
            ws = wb.active
            ws.title = "Munkaórák"
            wb.save(excel_path)

            headers = ["Dátum", "Nap", 
                      "Sima Munkanap\nKezdés", "Sima Munkanap\nVégzés", "Ledolgozott\nÓrák",
                      "Műhely\nKezdés", "Műhely\nVégzés", "Ledolgozott\nÓrák",
                      "Szabadság", "Betegszabadság\n(TP)"]
        
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for row in range(self.work_table.rowCount()):
                for col in range(self.work_table.columnCount()):
                    item = self.work_table.item(row, col)
                    if item:
                        ws.cell(row=row + 2, column=col + 1, value=item.text())
                        ws.cell(row=row + 2, column=col + 1).alignment = Alignment(horizontal='center', vertical='center')

            wb.save('munkaora_nyilvantartas.xlsx')

            if work_type == "Szabadság":
                self.parent.vacation_manager.updateVacationDays()

            QMessageBox.information(self.parent, "Siker", "Adatok mentve!")
        
        except Exception as e:
            print(f"Részletes hiba: {str(e)}")
            QMessageBox.warning(self.parent, "Hiba", f"Mentési hiba: {str(e)}")

    def saveWorkHoursToExcel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Munkaórák"

            headers = [
                "Dátum", "Nap", 
                "Sima Munkanap\nKezdés", "Sima Munkanap\nVégzés", "Ledolgozott\nÓrák",
                "Műhely\nKezdés", "Műhely\nVégzés", "Ledolgozott\nÓrák",
                "Szabadság", "Betegszabadság\n(TP)"
            ]

            header_font = Font(bold=True, color="000000")
            header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border

            data_font = Font(color="000000")
            data_alignment = Alignment(horizontal="center", vertical="center")
            data_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            for row in range(self.work_table.rowCount()):
                for col in [0, 1]:
                    cell = ws.cell(row=row + 2, column=col + 1)
                    item = self.work_table.item(row, col)
                    if item:
                        cell.value = item.text()
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = data_border

                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row + 2, column=col)
                    if not cell.value:
                        cell.value = ""
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = data_border

            column_widths = [12, 10, 15, 15, 12, 15, 15, 12, 10, 15]
            for i, width in enumerate(column_widths):
                ws.column_dimensions[chr(65 + i)].width = width

            for row in range(2, ws.max_row + 1):
                day_name = ws.cell(row=row, column=2).value
                if day_name in ["Szombat", "Vasárnap"]:
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            wb.save('munkaora_nyilvantartas.xlsx')
            QMessageBox.information(self.parent, "Siker", "Munkaórák mentve!")

        except Exception as e:
            QMessageBox.warning(self.parent, "Hiba", f"Munkaóra mentési hiba: {str(e)}")

    def updateWorkTable(self, data):
        for row in range(self.work_table.rowCount()):
            if self.work_table.item(row, 0).text() == data['date']:
                for col in range(2, self.work_table.columnCount()):
                    self.work_table.setItem(row, col, QTableWidgetItem(""))
            
                if data['type'] == "Sima munkanap":
                    self.work_table.setItem(row, 2, QTableWidgetItem(data['start_time']))
                    self.work_table.setItem(row, 3, QTableWidgetItem(data['end_time']))
                    self.work_table.setItem(row, 4, QTableWidgetItem(f"{data['hours']:.2f}"))
                elif data['type'] == "Műhely nap":
                    self.work_table.setItem(row, 5, QTableWidgetItem(data['start_time']))
                    self.work_table.setItem(row, 6, QTableWidgetItem(data['end_time']))
                    self.work_table.setItem(row, 7, QTableWidgetItem(f"{data['hours']:.2f}"))
                elif data['type'] == "Szabadság":
                    self.work_table.setItem(row, 8, QTableWidgetItem("1"))
                elif data['type'] == "Betegszabadság (TP)":
                    self.work_table.setItem(row, 9, QTableWidgetItem("1"))
            
                for col in range(2, self.work_table.columnCount()):
                    item = self.work_table.item(row, col)
                    if item:
                        item.setTextAlignment(Qt.AlignCenter)
                break

    def loadFromExcel(self):
        try:
            if not os.path.exists('munkaora_nyilvantartas.xlsx'):
                return
            
            wb = load_workbook('munkaora_nyilvantartas.xlsx')
            ws = wb.active
        
            for row in range(self.work_table.rowCount()):
                for col in range(2, self.work_table.columnCount()):
                    self.work_table.setItem(row, col, QTableWidgetItem(""))
        
            for excel_row in range(2, ws.max_row + 1):
                date = ws.cell(row=excel_row, column=1).value
                if not date:
                    continue
                
                for table_row in range(self.work_table.rowCount()):
                    if self.work_table.item(table_row, 0).text() == str(date):
                        for col in range(2, self.work_table.columnCount()):
                            value = ws.cell(row=excel_row, column=col).value
                            if value is not None:
                                item = QTableWidgetItem(str(value))
                                item.setTextAlignment(Qt.AlignCenter)
                                self.work_table.setItem(table_row, col-1, item)
                        break
                    
            QMessageBox.information(self.parent, "Siker", "Excel adatok betöltve!")
        
        except Exception as e:
            QMessageBox.warning(self.parent, "Hiba", f"Excel betöltési hiba: {str(e)}")