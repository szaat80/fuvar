from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QMessageBox
from PySide6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os
import json
from datetime import datetime

class DeliveryManager:
    def __init__(self, parent=None):
        self.parent = parent
        self.delivery_table = None
        self.stored_values = {}

    def setup_delivery_table(self, table):
        self.delivery_table = table
        self.setup_headers()
        
    def setup_headers(self):
        headers = ["Dátum"]
        headers.extend([f"Övezet {i}-{i+5}" for i in range(0, 45, 5)])
        self.delivery_table.setColumnCount(len(headers))
        self.delivery_table.setHorizontalHeaderLabels(headers)

    def handleM3Input(self):
        text = self.parent.m3_input.text().strip()
        try:
            text = text.replace(',', '.')
        
            if not text:
                QMessageBox.warning(self.parent, "Hiba", "Kérem adjon meg egy számot!")
                return
        
            value = float(text)
            if value < 0:
                QMessageBox.warning(self.parent, "Hiba", "Kérem pozitív számot adjon meg!")
                return
            
            current_zone = self.parent.km_combo.currentText()
            current_zone_col = self.getZoneColumn(current_zone)
            
            if not hasattr(self, 'stored_values'):
                self.stored_values = {}
            
            current_date = self.parent.date_edit.date().toString('yyyy-MM-dd')
            if current_date not in self.stored_values:
                self.stored_values[current_date] = {}
            
            if current_zone not in self.stored_values[current_date]:
                self.stored_values[current_date][current_zone] = []
        
            self.stored_values[current_date][current_zone].append(value)
            self.updateDeliveryTableWithStoredValues()
            self.parent.m3_input.clear()
            self.updateM3Sum(current_date, current_zone)
        
        except ValueError:
            QMessageBox.warning(self.parent, "Hiba", "Kérem számot adjon meg (pl.: 6.0 vagy 6,0)")

    def getZoneColumn(self, zone_text):
        try:
            start_km = int(zone_text.split(' ')[1].split('-')[0])
            return (start_km // 5) + 1
        except:
            return 0

    def updateM3Sum(self, current_date, current_zone):
        if current_date in self.stored_values and current_zone in self.stored_values[current_date]:
            values = self.stored_values[current_date][current_zone]
            values_text = " + ".join(f"{v:.1f}" for v in values)
            total = sum(values)
            self.parent.m3_sum.setText(f"({values_text}) = {total:.1f}")
        else:
            self.parent.m3_sum.setText("(0)")

    def updateDeliveryTableWithStoredValues(self):
        current_date = self.parent.date_edit.date().toString('yyyy-MM-dd')
    
        if current_date in self.stored_values:
            for row in range(self.delivery_table.rowCount()):
                if self.delivery_table.item(row, 0) and self.delivery_table.item(row, 0).text() == current_date:
                    for zone, values in self.stored_values[current_date].items():
                        col = self.getZoneColumn(zone)
                        if col > 0:
                            display_text = " + ".join(f"{v:.1f}" for v in values)
                            new_item = QTableWidgetItem(display_text)
                            new_item.setTextAlignment(Qt.AlignCenter)
                            self.delivery_table.setItem(row, col, new_item)

    def saveDeliveryData(self):
        try:
            current_driver = self.parent.driver_combo.currentText()
            month_dir = os.path.join('driver_records', current_driver, 
                                    f"{datetime.now().year}_{datetime.now().month:02d}")
            os.makedirs(month_dir, exist_ok=True)
            self.saveDeliveryToExcel(os.path.join(month_dir, 'deliveries.xlsx'))
            QMessageBox.information(self.parent, "Siker", "Fuvar adatok mentve!")
        
        except Exception as e:
            QMessageBox.warning(self.parent, "Hiba", f"Fuvar adat mentési hiba: {str(e)}")

    def saveDeliveryToExcel(self, filepath):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Fuvar adatok"
        
            headers = ["Dátum"]
            headers.extend([f"Övezet {i}-{i+5}" for i in range(0, 45, 5)])
        
            header_style = {
                'font': Font(bold=True),
                'fill': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            }
        
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                for key, value in header_style.items():
                    setattr(cell, key, value)
        
            for row in range(self.delivery_table.rowCount()):
                for col in range(len(headers)):
                    item = self.delivery_table.item(row, col)
                    if item:
                        cell = ws.cell(row=row+2, column=col+1, value=item.text())
                        cell.alignment = Alignment(horizontal="center", vertical="center")
        
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        
            wb.save(filepath)
        
        except Exception as e:
            QMessageBox.warning(self.parent, "Hiba", f"Fuvar adat Excel mentési hiba: {str(e)}")